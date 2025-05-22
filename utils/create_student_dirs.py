import os
import sys
from openpyxl import load_workbook

def get_initials(full_name):
    """Convert full name to initials (e.g., 'John Cena' -> 'JC')."""
    if not full_name:
        return "NA"
    return ''.join([part[0].upper() for part in full_name.strip().split() if part])

def create_directories_from_sheet(file_name, sheet_name):
    # Load the Excel workbook and sheet
    try:
        wb = load_workbook(file_name, data_only=True)
    except FileNotFoundError:
        print(f"Error: File '{file_name}' not found.")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' does not exist.")
        return

    sheet = wb[sheet_name]
    row_index = 7

    while True:
        student_id = sheet.cell(row=row_index, column=2).value  # Column B
        student_name = sheet.cell(row=row_index, column=4).value  # Column D

        if not student_id:
            break

        initials = get_initials(student_name)
        print(f"Processing: ID={student_id}, Name={student_name}, Initials={initials}")
        dir_name = f"{student_id}-{initials}"
        print(f"Directory name: {dir_name}")

        try:
            os.makedirs(dir_name, exist_ok=True)
            print(f"Created directory: {dir_name}")
        except Exception as e:
            print(f"Failed to create '{dir_name}': {e}")

        row_index += 1

if __name__ == "__main__":
    # Default Excel file name
    excel_file = "gen_10_list.xlsx"

    # Get sheet name from command-line argument
    if len(sys.argv) < 2:
        print("Usage: python create_student_dirs.py <SheetName>")
        sys.exit(1)

    sheet_name = sys.argv[1]
    create_directories_from_sheet(excel_file, sheet_name)
